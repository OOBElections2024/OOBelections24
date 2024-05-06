import streamlit as st
from openpyxl import load_workbook
import json
import pandas as pd
import plotly.express as px
import os
import base64

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")
data_file_path = "data/data.xlsx"

# Define paths to candidate images
candidate_images = {
    "candidate 1": "images/candidate1.png",
    "candidate 2": "images/candidate2.png",
    "candidate 3": "images/candidate3.png",
    "candidate 4": "images/candidate4.png",
    "candidate 5": "images/candidate5.png",
    "candidate 6": "images/candidate6.png",
    "candidate 7": "images/candidate7.png",
    "candidate 8": "images/candidate8.png",
}

# Define candidate names as variables
candidate1 = "Areeb Ahmed"
candidate2 = "Jaideep Anandan"
candidate3 = "Mithil Pramod"
candidate4 = "Sreehari Deepu Nair"
candidate5 = "Aaron George"
candidate6 = "Affan Mohammed"
candidate7 = "Avinash Prasad"
candidate8 = "Swamidas Nair"


def setup_session_state():
    if 'page' not in st.session_state:
        st.session_state.page = 'Home'


def landing_page():
    st.markdown("<h1 style='text-align: center; font-family: 'Roboto';'>OOB House Elections</h1>", unsafe_allow_html=True)

    app_mode = st.sidebar.radio("Go to", ["Jupiter", "Saturn", "Mars", "Neptune", "Admin"])
    st.session_state.page = app_mode.lower()

def display_jupiter_page():
    st.markdown("<h2 style='text-align: center; font-family: 'Roboto'; font-size: 24px;'>Jupiter House</h2>", unsafe_allow_html=True)
    
    with st.container():
        col1, col2 = st.columns(2)
        
        # Apply the style for the image borders (light shade of deep blue with thin border)
        image_border_color = """
            <style>
                .st-emotion-cache-1kyxreq.e115fcil2 {
                    display: flex;
                    flex-direction: column;
                    justify-content: center;
                    align-items: center;
                    text-align: center;
                }
                
                .st-emotion-cache-1o1eenq.e1nzilvr5 {
                    text-align: center;
                    font-size: 18px; /* Increase font size */
                }
            </style>
        """
        st.markdown(image_border_color, unsafe_allow_html=True)

        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/candidate1.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate1}</p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/image1.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/candidate2.png", width=130)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate2}</p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/image2.png", width=200)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)
    
    # Apply the style for the specified section (light shade of blue)
    section_color = """
        <style>
            [data-testid="stAppViewContainer"] {
                background-color: #b2cddb; /* Light shade of deep blue */
            }
        </style>
    """
    st.markdown(section_color, unsafe_allow_html=True)
    
    # Apply the style for the specified div tag class (greyish blue white)
    div_tag_color = """
        <style>
            .st-emotion-cache-r421ms.e10yg2by1 {
                background-color: #d8e6ed; /* Greyish blue white */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(div_tag_color, unsafe_allow_html=True)
    
    # Apply the style for textboxes (white color)
    textbox_color = """
        <style>
            input[type="text"], textarea {
                background-color: white; /* White color for textboxes */
                border-color: #000; /* Deep shade of blue */
                color: #073b4c; /* Deep shade of blue */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(textbox_color, unsafe_allow_html=True)
    
    # Apply the style for form (very pale shade of blue)
    form_color = """
        <style>
            .st-ax {
                background-color: #073b4c; /* Very pale shade of blue */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(form_color, unsafe_allow_html=True)
    
    vote_button_color = """
        <style>
            .css-19ih76x button {
                color: #073b4c !important;
                background-color: #d4e8f0 !important; /* Very pale shade of blue */
                border-color: #073b4c !important; /* Deep shade of blue */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(vote_button_color, unsafe_allow_html=True)
    
    with st.form(key='jupiter_form', clear_on_submit=True):
        candidate = st.radio("Select a candidate", (candidate1, candidate2))
        
        # Apply the style for form labels (increase font size)
        label_style = """
            <style>
                p {
                    font-size: 18px; /* Increase font size for form labels */
                    font-weight: bold;
                }
                .st-emotion-cache-l9bjmx e1nzilvr5 {
                    font-weight: bold;
                }
            </style>
        """
        st.markdown(label_style, unsafe_allow_html=True)
        
        student_id = st.text_input("Admission No. ID (The last 5 digits on your Bus card No.) (Eg: 57122)", "", max_chars=5)
        grade_section = st.text_input("Grade and Section (Eg: 9A)", "", max_chars=3)
        if st.form_submit_button("Vote"):
            if student_id and grade_section:
                vote(candidate, student_id, grade_section, "Jupiter")
            else:
                st.error("Please enter ID, Grade, and Section.")

def display_saturn_page():
    st.markdown("<h2 style='text-align: center; font-family: 'Roboto'; font-size: 24px;'>Saturn House</h2>", unsafe_allow_html=True)
    
    # Apply the style for the background (light yellow)
    background_color = """
        <style>
            [data-testid="stAppViewContainer"] {
                background-color: #f7c945; /* Light yellow */
            }
        </style>
    """
    st.markdown(background_color, unsafe_allow_html=True)
    
    with st.container():
        col1, col2 = st.columns(2)
        
        # Apply the style for the image borders (light shade of yellow with thin border)
        image_border_color = """
            <style>
                .st-emotion-cache-1kyxreq.e115fcil2 {
                    display: flex;
                    flex-direction: column;
                    justify-content: center;
                    align-items: center;
                    text-align: center;
                }
                
                .st-emotion-cache-1o1eenq.e1nzilvr5 {
                    text-align: center;
                    font-size: 18px; /* Increase font size */
                }
            </style>
        """
        st.markdown(image_border_color, unsafe_allow_html=True)
        
        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/candidate3.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate3}</p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/image3.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/candidate4.png", width=220)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate4}</p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/image4.png", width=200)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)
    
    # Apply the style for the specified section (greyish yellow)
    section_color = """
        <style>
            .st-emotion-cache-r421ms.e10yg2by1 {
                background-color: #fada7f; /* Greyish yellow */
            }
        </style>
    """
    st.markdown(section_color, unsafe_allow_html=True)
    
    # Apply the style for textboxes (white color)
    textbox_color = """
        <style>
            input[type="text"], textarea {
                background-color: #ededed; /* White color for textboxes */
                border-color: #000; /* Deep shade of blue */
                color: #073b4c; /* Deep shade of blue */
            }
        </style>
    """
    st.markdown(textbox_color, unsafe_allow_html=True)
    
    # Apply the style for form (slightly darker shade of yellow)
    form_color = """
        <style>
            .st-ax {
                background-color: #dcc88c; /* Slightly darker shade of yellow */
                color: #c9ad61; /* Slightly darker shade of yellow */
            }
        </style>
    """
    st.markdown(form_color, unsafe_allow_html=True)
    
    vote_button_color = """
        <style>
            .css-19ih76x button {
                color: #b0a402 !important;
                background-color: #b0a402 !important; /* Dark shade of yellow */
                border-color: #c9ad61 !important; /* Dark shade of yellow */
            }
        </style>
    """
    st.markdown(vote_button_color, unsafe_allow_html=True)
    
    with st.form(key='saturn_form', clear_on_submit=True):
        candidate = st.radio("Select a candidate", (candidate3, candidate4))
        
        # Apply the style for form labels (increase font size)
        label_style = """
            <style>
                p {
                    font-size: 18px; /* Increase font size for form labels */
                    font-weight: bold;
                }
                .st-emotion-cache-l9bjmx e1nzilvr5 {
                    font-weight: bold;
                }
            </style>
        """
        st.markdown(label_style, unsafe_allow_html=True)
        
        student_id = st.text_input("Admission No. ID (The last 5 digits on your Bus card No.) (Eg: 57122)", "", max_chars=5)
        grade_section = st.text_input("Grade and Section (Eg: 9A)", "")
        if st.form_submit_button("Vote"):
            if student_id and grade_section:
                vote(candidate, student_id, grade_section, "Saturn")
            else:
                st.error("Please enter ID, Grade, and Section.")

def display_mars_page():
    st.markdown("<h2 style='text-align: center; font-family: 'Roboto'; font-size: 24px;'>Mars House</h2>", unsafe_allow_html=True)
    
    # Apply the style for the background (light red)
    background_color = """
        <style>
            [data-testid="stAppViewContainer"] {
                background-color: #f57373; /* Light red */
            }
        </style>
    """
    st.markdown(background_color, unsafe_allow_html=True)
    
    with st.container():
        col1, col2 = st.columns(2)
        
        # Apply the style for the image borders (light shade of red with thin border)
        image_border_color = """
            <style>
                .st-emotion-cache-1kyxreq.e115fcil2 {
                    display: flex;
                    flex-direction: column;
                    justify-content: center;
                    align-items: center;
                    text-align: center;
                }
                
                .st-emotion-cache-1o1eenq.e1nzilvr5 {
                    text-align: center;
                    font-size: 18px; /* Increase font size */
                }
            </style>
        """
        st.markdown(image_border_color, unsafe_allow_html=True)
        
        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/candidate5.png", width=230)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate5}</p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/image5.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/candidate6.png", width=200)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate6}</p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/image6.png", width=200)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)
    
    # Apply the style for the specified section (light red)
    section_color = """
        <style>
            .st-emotion-cache-r421ms.e10yg2by1 {
                background-color: #ffcccc; /* Light red */
            }
        </style>
    """
    st.markdown(section_color, unsafe_allow_html=True)
    
    # Apply the style for textboxes (light red)
    textbox_color = """
        <style>
            input[type="text"], textarea {
                background-color: white; /* White color for textboxes */
                border-color: #073b4c; /* Deep shade of blue */
                color: #073b4c; /* Deep shade of blue */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(textbox_color, unsafe_allow_html=True)
    
    # Apply the style for form (slightly darker shade of red)
    form_color = """
        <style>
            .st-ax {
                background-color: #ff9999; /* Lighter shade of red */
                color: #ff9999; /* Lighter shade of red */
            }
        </style>
    """
    st.markdown(form_color, unsafe_allow_html=True)
    
    vote_button_color = """
        <style>
            .css-19ih76x button {
                color: #ff9999 !important;
                background-color: #ff6666 !important; /* Light shade of red */
                border-color: #ff6666 !important; /* Light shade of red */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(vote_button_color, unsafe_allow_html=True)
    
    with st.form(key='mars_form', clear_on_submit=True):
        candidate = st.radio("Select a candidate", (candidate5, candidate6))
        
        # Apply the style for form labels (increase font size)
        label_style = """
            <style>
                p {
                    font-size: 18px; /* Increase font size for form labels */
                    font-weight: bold;
                }
                .st-emotion-cache-l9bjmx e1nzilvr5 {
                    font-weight: bold;
                }
            </style>
        """
        st.markdown(label_style, unsafe_allow_html=True)
        
        student_id = st.text_input("Admission No. ID (The last 5 digits on your Bus card No.) (Eg: 57122)")
        grade_section = st.text_input("Grade and Section (Eg: 9A)", "")
        if st.form_submit_button("Vote"):
            if student_id and grade_section:
                vote(candidate, student_id, grade_section, "Mars")
            else:
                st.error("Please enter ID, Grade, and Section.")

def display_neptune_page():
    st.markdown("<h2 style='text-align: center; font-family: 'Roboto'; font-size: 24px;'>Neptune House</h2>", unsafe_allow_html=True)
    
    # Apply the style for the background (light shade of forest green)
    background_color = """
        <style>
            [data-testid="stAppViewContainer"] {
                background-color: #419641; /* Light shade of forest green */
            }
        </style>
    """
    st.markdown(background_color, unsafe_allow_html=True)
    
    with st.container():
        col1, col2 = st.columns(2)
        
        # Apply the style for the image borders (light shade of forest green with thin border)
        image_border_color = """
            <style>
                .st-emotion-cache-1kyxreq.e115fcil2 {
                    display: flex;
                    flex-direction: column;
                    justify-content: center;
                    align-items: center;
                    text-align: center;
                }
                
                .st-emotion-cache-1o1eenq.e1nzilvr5 {
                    text-align: center;
                    font-size: 18px; /* Increase font size */
                }
            </style>
        """
        st.markdown(image_border_color, unsafe_allow_html=True)
        
        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/candidate7.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate7}</p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col1.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col1.image("images/image7.png", width=200)
        col1.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'></p>", unsafe_allow_html=True)
        col1.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/candidate8.png", width=190)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'>{candidate8}</p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)

        col2.markdown("<div class='st-emotion-cache-1kyxreq e115fcil2'>", unsafe_allow_html=True)
        col2.image("images/image8.png", width=200)
        col2.markdown(f"<p class='st-emotion-cache-1o1eenq e1nzilvr5' style='font-size: 18px;'><p>", unsafe_allow_html=True)
        col2.markdown("</div>", unsafe_allow_html=True)
    
    section_color = """
        <style>
            .st-emotion-cache-r421ms.e10yg2by1 {
                background-color: #badbba; /* Light shade of forest green */
            }
        </style>
    """
    st.markdown(section_color, unsafe_allow_html=True)
    
    textbox_color = """
        <style>
            input[type="text"], textarea {
                background-color: white; /* White color for textboxes */
                border-color: #073b4c; /* Deep shade of blue */
                color: #073b4c; /* Deep shade of blue */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(textbox_color, unsafe_allow_html=True)
    
    # Apply the style for form (slightly darker shade of forest green)
    form_color = """
        <style>
            .st-ax {
                background-color: #669966; /* Slightly darker shade of forest green */
                color: #669966; /* Slightly darker shade of forest green */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(form_color, unsafe_allow_html=True)
    
    vote_button_color = """
        <style>
            .css-19ih76x button {
                color: #669966 !important;
                background-color: #669966 !important; /* Light shade of forest green */
                border-color: #669966 !important; /* Light shade of forest green */
                font-size: 18px; /* Increase font size */
            }
        </style>
    """
    st.markdown(vote_button_color, unsafe_allow_html=True)
    
    with st.form(key='neptune_form', clear_on_submit=True):
        candidate = st.radio("Select a candidate", (candidate7, candidate8))
        
        # Apply the style for form labels (increase font size)
        label_style = """
            <style>
                p {
                    font-size: 18px; /* Increase font size for form labels */
                    font-weight: bold;
                }
                .st-emotion-cache-l9bjmx e1nzilvr5 {
                    font-weight: bold;
                }
            </style>
        """
        st.markdown(label_style, unsafe_allow_html=True)
        
        student_id = st.text_input("Admission No. ID (The last 5 digits on your Bus card No.) (Eg: 57122)")
        grade_section = st.text_input("Grade and Section (Eg: 9A)", "")
        if st.form_submit_button("Vote"):
            if student_id and grade_section:
                vote(candidate, student_id, grade_section, "Neptune")
            else:
                st.error("Please enter ID, Grade, and Section.")

def vote(candidate, student_id, grade_section, house):
    try:
        workbook = load_workbook(filename=data_file_path)
        sheet = workbook.active

        id_found = False
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
            if str(row[0]) == student_id and row[2].lower() == grade_section.lower() and row[3].lower() == house.lower():
                id_found = True

                voting_status = row[4]
                if voting_status == 'no':
                    sheet.cell(row=row_num, column=5, value='Yes')
                    workbook.save(data_file_path)
                    sheet.cell(row=row_num, column=6, value=candidate)
                    workbook.save(data_file_path)

                    st.success("Thank you for voting!")

                    update_vote_count(house.lower(), candidate.lower())
                    return
                else:
                    st.error("You have already voted.")
                    return
        else:
            st.error("Incorrect information or house.")
            
    except Exception as e:
        st.error(f"Error: {str(e)}")

def update_vote_count(house, candidate):
    with open("data/vote_counts.json", "r") as file:
        vote_counts = json.load(file)

    if house in vote_counts:
        if candidate in vote_counts[house]:
            vote_counts[house][candidate] += 1
        else:
            vote_counts[house][candidate] = 1
    else:
        vote_counts[house] = {candidate: 1}

    with open("data/vote_counts.json", "w") as file:
        json.dump(vote_counts, file, indent=2)

def display_class_statistics(class_name):
    try:
        workbook = load_workbook(filename=data_file_path)
        sheet = workbook.active

        voted_students = []
        not_voted_students = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == class_name:
                if row[4].lower() == "yes":  # Check voting status in column E (index 4)
                    voted_students.append(row[1])  
                elif row[4].lower() == "no":  # Check voting status in column E (index 4)
                    not_voted_students.append(row[1])  

        fig = px.pie(
            values=[len(voted_students), len(not_voted_students)],
            names=["Voted", "Not Voted"],
            title=f"Voting Status for Class {class_name.upper()}",
            color_discrete_sequence=["#1f77b4", "#ff7f0e"]
        )
        st.plotly_chart(fig)

        st.subheader("Names of Students Who Have Not Voted:")
        if not_voted_students:
            for student in not_voted_students:
                st.write(student)
        else:
            st.write("No students have not voted.")
    except Exception as e:
        st.error(f"Error: {str(e)}")

def display_house_statistics(house_name):
    try:
        workbook = load_workbook(filename=data_file_path)
        sheet = workbook.active

        voted_students = []
        not_voted_students = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] and house_name and row[3].lower() == house_name.lower():
                if row[4] and row[4].lower() == "yes":
                    voted_students.append(row[1])  # Change row index to 1 for name
                elif row[4] and row[4].lower() == "no":
                    not_voted_students.append(row[1])  # Change row index to 1 for name

        fig = px.pie(
            values=[len(voted_students), len(not_voted_students)],
            names=["Voted", "Not Voted"],
            title=f"Voting Status for House {house_name.capitalize()}",
            color_discrete_sequence=["#1f77b4", "#ff7f0e"]
        )
        st.plotly_chart(fig)

        st.subheader("Names of Students Who Have Not Voted:")
        if not_voted_students:
            for student in not_voted_students:
                st.write(student)
        else:
            st.write("All students have voted.")
    except Exception as e:
        st.error(f"Error: {str(e)}")

def show_results():
    with open("data/vote_counts.json", "r") as file:
        vote_counts = json.load(file)

    for house, candidates in vote_counts.items():
        with st.expander(f"Results for {house.capitalize()} House"):
            # Find the winning candidate
            winning_candidate = max(candidates, key=candidates.get)
            # Find the losing candidate
            losing_candidate = min(candidates, key=candidates.get)

            # Display the results
            if winning_candidate and losing_candidate:
                # Use columns for layout
                col1, col2 = st.columns([3, 1])
                
                # Header with winning candidate's name
                with col1:
                    st.subheader(winning_candidate)
                
                # Subheader with title
                with col1:
                    st.write(f"Vice Captain - {house.capitalize()}")
                    st.write("")  # Add empty line for spacing
                    st.write("") 
                    st.write("") 
                    st.write("")  
                    st.write("")  
                    st.write("")  
                
                # Floating image to the right
                with col2:
                    if winning_candidate in candidate_images:
                        img_path = candidate_images[winning_candidate]
                        st.image(img_path, width=150)  # Set a fixed width for the image
                    else:
                        st.write("Image not found.")
                
                # Losing candidate
                with col1:
                    st.subheader(losing_candidate)
                with col1:
                    st.write(f"Senior Prefect - {house.capitalize()}")
                    st.write("")  # Add empty line for spacing
                    st.write("")
                    st.write("")
                    st.write("")
                    st.write("")
                    st.write("")
                with col2:
                    if losing_candidate in candidate_images:
                        img_path = candidate_images[losing_candidate]
                        st.image(img_path, width=150)
                    else:
                        st.write("Image not found.")
            else:
                st.write("Error: No winning or losing candidate found.")
                
def add_person_details(person_id, person_name, person_class, person_house):
    try:
        workbook = load_workbook(filename=data_file_path)
        sheet = workbook.active

        # Check if the person_id already exists
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == person_id:
                st.error("Person with this ID already exists.")
                return

        # Find the first empty row in column A
        empty_row = 1
        while sheet.cell(row=empty_row, column=1).value is not None:
            empty_row += 1

        # Append person details to the topmost empty row
        new_row = [person_id, person_name, person_class, person_house, "no"]
        for col, value in enumerate(new_row, start=1):
            sheet.cell(row=empty_row, column=col, value=value)

        # Save the workbook
        workbook.save(data_file_path)
        st.success("Person details added successfully!")
    except Exception as e:
        st.error(f"Error: {str(e)}")

def remove_student(student_data):
    try:
        # Ensure the student_data format is valid
        if len(student_data) < 1:
            st.error("Invalid student data format.")
            return

        workbook = load_workbook(filename=data_file_path)
        sheet = workbook.active

        # Find the row matching the student data
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == str(student_data[0]):  # Compare student ID
                st.write("DEBUG: Found matching row:", row)  # Debugging statement

                # Get the voted candidate and house
                candidate_name = row[5]  # Candidate name in column F
                house = row[3]

                # Remove the row
                sheet.delete_rows(row_num, 1)
                workbook.save(data_file_path)
                st.success("Student removed successfully!")

                if candidate_name:  # If the student has voted
                    # Decrement the vote count of the voted candidate
                    with open("data/vote_counts.json", "r") as file:
                        vote_counts = json.load(file)

                    if house.lower() in vote_counts and candidate_name.lower() in vote_counts[house.lower()]:
                        vote_counts[house.lower()][candidate_name.lower()] -= 1
                        with open("data/vote_counts.json", "w") as file:
                            json.dump(vote_counts, file, indent=2)
                        st.write(f"DEBUG: Vote count for {candidate_name} decremented")  # Debugging message

                return

        st.error("Student not found.")
    except Exception as e:
        st.error(f"Error: {str(e)}")

def display_remove_student():
    st.subheader("Remove a Student")

    # Select search criteria
    search_criteria = st.radio("Search by", ["ID", "Name", "Class", "House"])

    # Text input for search query
    search_query = st.text_input(f"Enter {search_criteria}:")

    # Find students matching the search query
    workbook = load_workbook(filename=data_file_path)
    sheet = workbook.active
    students = []

    # Handle each search criteria
    if search_criteria == "ID":
        students = [str(row[0]) for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
    elif search_criteria == "Name":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]
    elif search_criteria == "Class":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]
    elif search_criteria == "House":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]

    # Filter students based on search query
    filtered_students = [student for student in students if student and search_query.lower() in str(student).lower()]

    # Dropdown to select a student
    student_to_remove = st.selectbox("Select Student to Remove", filtered_students, index=0 if filtered_students else None)

    # Button to remove the selected student
    if st.button("Remove Student") and student_to_remove:
        student_data = student_to_remove.split(", ")
        remove_student(student_data)
        
def download_csv():
    # Load the Excel file
    df = pd.read_excel(data_file_path)
    
    # Create a CSV file in memory
    csv_data = df.to_csv(index=False)
    b64 = base64.b64encode(csv_data.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="data.csv">Download CSV File</a>'
    st.markdown(href, unsafe_allow_html=True)

def download_json():
    # Load vote counts JSON
    with open("data/vote_counts.json", "r") as file:
        vote_counts = file.read()
    
    # Create a JSON file in memory
    b64 = base64.b64encode(vote_counts.encode()).decode()
    href = f'<a href="data:file/json;base64,{b64}" download="vote_counts.json">Download JSON File</a>'
    st.markdown(href, unsafe_allow_html=True)

def search_student():
    st.subheader("Search Student")

    # Select search criteria
    search_criteria = st.radio("Search by", ["ID", "Name", "Class", "House"], key="search_criteria")

    # Text input for search query
    search_query = st.text_input(f"Enter {search_criteria}:", key=f"search_input_{search_criteria}")

    # Find students matching the search query
    workbook = load_workbook(filename=data_file_path)
    sheet = workbook.active
    students = []

    # Handle each search criteria
    if search_criteria == "ID":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]
    elif search_criteria == "Name":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]
    elif search_criteria == "Class":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]
    elif search_criteria == "House":
        students = [f"{row[0]}, {row[1]}, {row[2]}, {row[3]}" for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]

    # Filter students based on search query
    filtered_students = [student.split(", ") for student in students if student and search_query.lower() in str(student).lower()]

    # Dropdown to select a student
    selected_student = st.selectbox("Select Student", filtered_students, index=0 if filtered_students else None, key="select_student")

    # Print debug information
    if selected_student:
        st.write("Selected Student:")
        st.write(selected_student)

def display_admin_page():
    # Store password in session state if not already stored
    if 'admin_password' not in st.session_state:
        st.session_state.admin_password = None

    # Check if password is already stored or new password is entered
    if st.session_state.admin_password is None:
        entered_password = st.text_input("Enter Admin Password:", type="password")
    else:
        entered_password = st.session_state.admin_password

    if entered_password == ADMIN_PASSWORD:
        # Store the validated password in session state
        st.session_state.admin_password = entered_password

        # Display admin page content
        st.title("Admin Page")

        # Add a refresh button to update the admin page content
        refresh_button = st.button("Refresh Admin Page")

        if refresh_button:
            # Rerun the page to refresh the content
            st.rerun()

        st.header("Total Votes by House and Candidate")

        with open("data/vote_counts.json", "r") as file:
            vote_counts = json.load(file)

        color_sequences = {
            "jupiter": ["#aec7e8", "#1f77b4"],  # Light and dark blue
            "saturn": ["#ffbb78", "#ff7f0e"],   # Light and dark yellow
            "mars": ["#ff9896", "#d62728"],     # Light and dark red
            "neptune": ["#98df8a", "#2ca02c"],  # Light and dark green
        }

        for house, candidates in vote_counts.items():
            with st.expander(f"{house.capitalize()} House"):
                st.subheader(f"{house.capitalize()} House")
                fig = px.pie(
                    names=list(candidates.keys()),
                    values=list(candidates.values()),
                    title=f"Total Votes for {house.capitalize()} House",
                    color_discrete_sequence=color_sequences.get(house.lower(), "Plotly")
                )
                st.plotly_chart(fig)

                st.write("Vote Counts:")
                df = pd.DataFrame.from_dict(candidates, orient='index', columns=['Votes'])
                st.write(df)

        st.header("Statistics")
        selection = st.radio("Select Statistics", ["Class Statistics", "House Statistics"])
        if selection == "Class Statistics":
            class_name = st.text_input("Enter Class Name (e.g., 9A):")
            if class_name:
                display_class_statistics(class_name)
        elif selection == "House Statistics":
            house_name = st.radio("Select House", ["Jupiter", "Saturn", "Mars", "Neptune"])
            display_house_statistics(house_name)
            
        with st.expander("Search Student"):
            search_student()
            
        with st.expander("Add Person Details"):    
            st.subheader("Add Person Details")

            # Text input for ID, name, class, and house
            person_id = st.text_input("ID:")
            person_name = st.text_input("Name:")
            person_class = st.text_input("Class:")
            person_house = st.text_input("House:")

            # Button to add person details to the Excel file
            if st.button("Add Person"):
                if person_id and person_name and person_class and person_house:
                    add_person_details(person_id, person_name, person_class, person_house)
                else:
                    st.error("Please fill in all fields.")
                    
        with st.expander("Remove a Student"):
            display_remove_student()

        if st.button("Download CSV"):
            download_csv()
        if st.button("Download JSON"):
            download_json()
        if st.button("Get Results"):
            show_results()
        
    else:
        st.error("Incorrect admin password. Access denied.")


def switch_page():
    if st.session_state.page == 'jupiter':
        display_jupiter_page()
    elif st.session_state.page == 'saturn':
        display_saturn_page()
    elif st.session_state.page == 'mars':
        display_mars_page()
    elif st.session_state.page == 'neptune':
        display_neptune_page()
    elif st.session_state.page == 'admin':
        display_admin_page()


def main():
    setup_session_state()
    landing_page()
    switch_page()


if __name__ == "__main__":
    main()
